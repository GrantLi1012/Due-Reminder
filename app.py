import openpyxl
import tkinter
from tkinter import *

#window
root = Tk()
root.title("Deadline Manager")

#variables
dueIn = 7


#method
def change_due_in(num):
    global dueIn
    dueIn = num


#widgets
infoLabel = Label(root, text="Due in the next" + dueIn + "days:").grid(row=1, column=1)


wb = openpyxl.load_workbook("C:/Users/30494/PycharmProjects/Due-Reminder/task.xlsx", read_only=False, keep_vba=True)
sheet = wb.active
cell = sheet.cell(2, 2)
print(cell.value)
cell1 = sheet.cell(6, 6)
cell1.value = 2333

wb.save("C:/Users/30494/PycharmProjects/Due-Reminder/test/task.xlsx")

print(sheet.cell(6, 6).value)
