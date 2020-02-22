import openpyxl

wb = openpyxl.load_workbook("C:/Users/30494/PycharmProjects/Due-Reminder/test/tasks.xlsx", read_only=False, keep_vba=True)
sheet = wb.active
cell = sheet.cell(2, 2)
print(cell.value)
sheet['F6'] = 23333

wb.save("C:/Users/30494/PycharmProjects/Due-Reminder/test/tasks.xlsx")

print(sheet.cell(6, 6).value)
